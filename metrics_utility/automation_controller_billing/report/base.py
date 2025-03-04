######################################
# Code for building the spreadsheet
######################################
import os
import json

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from metrics_utility.automation_controller_billing.helpers import merge_json_sets, merge_arrays


class Base:
    BLACK_COLOR_HEX = "00000000"
    WHITE_COLOR_HEX = "00FFFFFF"
    BLUE_COLOR_HEX = "000000FF"
    RED_COLOR_HEX = "FF0000"
    LIGHT_BLUE_COLOR_HEX = "d4eaf3"
    GREEN_COLOR_HEX = "92d050"
    FONT = "Arial"
    PRICE_FORMAT = '$#,##0.00'

    @staticmethod
    def optional_report_sheets():
        return os.environ.get(
            'METRICS_UTILITY_OPTIONAL_CCSP_REPORT_SHEETS',
            'ccsp_summary,managed_nodes,usage_by_organizations,usage_by_collections,usage_by_roles,'\
            'usage_by_modules').split(",")

    def convert_cell(self, cell):
        # If the cell is a dictionary, convert each set value to a sorted list, then dump as a JSON string.
        if isinstance(cell, dict):
            new_cell = {k: sorted(list(v)) if isinstance(v, set) else v for k, v in cell.items()}
            return json.dumps(new_cell)
        # If the cell itself is a set, convert it to a sorted list and then to a JSON string.
        elif isinstance(cell, set):
            return json.dumps(sorted(list(cell)))
        # If the cell is a list, convert any set elements inside to sorted lists and dump as a JSON string.
        elif isinstance(cell, list):
            new_cell = [sorted(list(item)) if isinstance(item, set) else item for item in cell]
            return json.dumps(new_cell)
        # Otherwise, return the cell unchanged.
        return cell

    def _fix_event_host_names(self, mapping_dataframe, destination_dataframe):
        if destination_dataframe is None:
            return None

        def concatenate_columns_mapping(row):
            return f"{row['original_host_name']}__{str(row['install_uuid'])}__{str(row['job_remote_id'])}"

        def concatenate_columns_destination(row):
            return f"{row['host_name']}__{str(row['install_uuid'])}__{str(row['job_remote_id'])}"

        # Apply the function to each row of the DataFrame
        mapping_dataframe['host_composite_id'] = mapping_dataframe.apply(concatenate_columns_mapping, axis=1)
        mapping_dataframe = mapping_dataframe.set_index("host_composite_id")
        mapping_dataframe = mapping_dataframe["host_name"].astype(str).to_dict()

        def apply_mapping(row):
            return mapping_dataframe.get(f"{row['host_name']}__{str(row['install_uuid'])}__{row['job_remote_id']}", row['host_name'])

        destination_dataframe['host_name'] = destination_dataframe.apply(apply_mapping, axis=1)
        destination_dataframe['host_composite_id'] = destination_dataframe.apply(concatenate_columns_destination, axis=1)

        return destination_dataframe


    def _build_data_section_usage_by_node_with_org_details(self, current_row, ws, dataframe, mode=None, managed_node_type=None):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Rename the columns based on the template
        ccsp_report_dataframe = (
            dataframe.groupby('host_name', dropna=False)
            .agg(
                organizations=('organization_name', 'nunique'),
                host_runs=('host_name', 'count'),
                task_runs=('task_runs', 'sum'),
                first_automation=('first_automation', 'min'),
                last_automation=('last_automation', 'max')
            )
        )
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()
        columns = [
            'host_name',
            'organizations',
            'host_runs',
            'task_runs',
            'first_automation',
            'last_automation',
        ]
        if mode == "by_organization":
            # Filter some columns out based on mode
            columns = [col for col in columns if col not in ['organizations']]

        ccsp_report_dataframe = ccsp_report_dataframe.reindex(
            columns=columns
        )

        # Create dataframe with hostname and orgs as columns, having last automation for each host
        pivoted_dataframe = dataframe.pivot_table(
            index='host_name',
            columns='organization_name',
            values='last_automation',
            aggfunc='max'  # You can use 'max', 'min', 'mean', etc., depending on your needs
        )

        # Set index on host_name for join
        ccsp_report_dataframe.set_index('host_name', inplace=True)

        # Join the list of orgs to the pivoted_dataframe having org last updated as columns
        ccsp_report_dataframe = ccsp_report_dataframe.join(pivoted_dataframe, how='left')
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        labels = {
            "host_name": "Host name",
            "organizations": "Automated by\norganizations",
            "host_runs": "Job runs",  # Job runs is the same as host_runs, Non-unique managed nodes automated
            "task_runs": "Number of task\nruns",
            'first_automation': "First\nautomation",
            'last_automation': "Last\nautomation",
        }
        labels = {k:v for k, v in labels.items() if k in columns}
        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns=labels
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter


    def _build_data_section_usage_by_job(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        dataframe['job_remote_id_install_uuid'] = list(zip(dataframe['job_remote_id'], dataframe['install_uuid']))

        # Rename the columns based on the template
        ccsp_report_dataframe = (
            dataframe.groupby(['organization_name', 'job_template_name'], dropna=False)
            .agg(
                job_runs=('job_remote_id_install_uuid', 'nunique'),
                host_runs_unique=('host_name', 'nunique'),
                host_runs=('host_name', 'count'),
                task_runs=('task_runs', 'sum'),
                first_run=('job_created', 'min'),
                last_run=('job_created', 'max'),
            )
        )
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()
        ccsp_report_dataframe = ccsp_report_dataframe.reindex(
            columns=[
                'job_template_name',
                'organization_name',
                'job_runs',
                'host_runs_unique',
                'host_runs',
                'task_runs',
                'first_run',
                'last_run'
            ]
        )

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "job_template_name": "Job template\nname",
                "organization_name": "Organization\nname",
                "job_runs": "Job runs",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
                "first_run": "First\nrun",
                "last_run": "Last\nrun",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_scope(self, current_row, ws, dataframe, mode=None):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        ccsp_report_dataframe = dataframe

        # Convert arrays and dict fields into string, so they can be rendered into xlsx
        for col in ['organizations', 'inventories', 'canonical_facts', 'facts']:
            ccsp_report_dataframe[col] = ccsp_report_dataframe[col].apply(self.convert_cell)

        # We're not showing cluster/install_uuid until we support multi-cluster view officially
        del ccsp_report_dataframe['install_uuid']

        columns = [
            'host_name',
            'last_automation',
            'organizations',
            'inventories',
            'canonical_facts',
            'facts',
        ]

        labels = {
            "host_name": "Host name",
            "last_automation": "Last\nAutomation",
            "organizations": "Organizations",
            "inventories": "Inventories",  # Job runs is the same as host_runs, Non-unique managed nodes automated
            "canonical_facts": "Canonical Facts",
            'facts': "Facts",
        }
        labels = {k:v for k, v in labels.items() if k in columns}
        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns=labels
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter


    def _build_data_section_usage_by_node(self, current_row, ws, dataframe, mode=None, managed_node_type=None):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        agg_dict = {
            "organizations": ('organization_name', 'nunique'),
            "host_runs": ('host_name', 'count'),
            "task_runs": ('task_runs', 'sum'),
            "first_automation": ('first_automation', 'min'),
            "last_automation": ('last_automation', 'max'),
            "manage_node_types": ('manage_node_types', lambda x: merge_arrays(x)),
            "events": ('events', lambda x: merge_arrays(x)),
            "canonical_facts": ('canonical_facts', lambda x: merge_json_sets(x)),
            "facts": ('facts', lambda x: merge_json_sets(x)),
        }

        # Now pass this dictionary into .agg()
        ccsp_report_dataframe = (
            dataframe
            .groupby("host_name", dropna=False)
            .agg(**agg_dict)
        )

        # Convert arrays and dict fields into string, so they can be rendered into xlsx
        for col in ['manage_node_types', 'events', 'canonical_facts', 'facts']:
            ccsp_report_dataframe[col] = ccsp_report_dataframe[col].apply(self.convert_cell)

        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()
        columns = [
            'host_name',
            'organizations',
            'host_runs',
            'task_runs',
            'first_automation',
            'last_automation',
        ]
        if managed_node_type == "indirect":
            columns += ['manage_node_types', 'canonical_facts', 'facts', 'events']

        if mode == "by_organization":
            # Filter some columns out based on mode
            columns = [col for col in columns if col not in ['organizations']]
        ccsp_report_dataframe = ccsp_report_dataframe.reindex(
            columns=columns
        )

        labels = {
            "host_name": "Host name",
            "organizations": "Automated by\norganizations",
            "host_runs": "Job runs",  # Job runs is the same as host_runs, Non-unique managed nodes automated
            "task_runs": "Number of task\nruns",
            'first_automation': "First\nautomation",
            'last_automation': "Last\nautomation",
        }
        if managed_node_type == "indirect":
            labels.update({
                "manage_node_types": "Manage\nNode\nTypes",
                "canonical_facts": "Canonical\nFacts",
                "facts": "Facts",
                "events": "Events",
            })

        labels = {k:v for k, v in labels.items() if k in columns}
        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns=labels
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_collections(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Take the content explorer dataframe and extract specific group by
        ccsp_report_dataframe = dataframe.groupby(
            ["collection_name"], dropna=False
        ).agg(
            host_runs_unique=('host_name', 'nunique'),
            host_runs=('host_composite_id', 'nunique'),
            task_runs=('task_runs', 'sum'),
            duration=('duration', "sum"))

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "collection_name": "Collection name",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
                "duration": "Duration of task\nruns [seconds]",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_roles(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        # Take the content explorer dataframe and extract specific group by
        ccsp_report_dataframe = dataframe.groupby(
            ["role_name"], dropna=False
        ).agg(
            host_runs_unique=('host_name', 'nunique'),
            host_runs=('host_composite_id', 'nunique'),
            task_runs=('task_runs', 'sum'),
            duration=('duration', "sum"))

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "role_name": "Role name",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
                "duration": "Duration of task\nruns [seconds]",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                # cell.border = dotted_border

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter

    def _build_data_section_usage_by_modules(self, current_row, ws, dataframe):
        for key, value in self.config['data_column_widths'].items():
            ws.column_dimensions[get_column_letter(key)].width = value

        header_font = Font(name=self.FONT,
                           size=10,
                           color=self.BLACK_COLOR_HEX,
                           bold=True)
        value_font = Font(name=self.FONT,
                          size=10,
                          color=self.BLACK_COLOR_HEX)

        ccsp_report_dataframe = dataframe.groupby(
            ["module_name"], dropna=False
        ).agg(
            host_runs_unique=('host_name', 'nunique'),
            host_runs=('host_composite_id', 'nunique'),
            task_runs=('task_runs', 'sum'),
            duration=('duration', "sum"))

        # Rename the columns based on the template
        ccsp_report_dataframe = ccsp_report_dataframe.reset_index()

        ccsp_report_dataframe = ccsp_report_dataframe.rename(
            columns={
                "module_name": "Module name",
                "host_runs_unique": "Unique managed nodes\nautomated",
                "host_runs": "Non-unique managed\nnodes automated",
                "task_runs": "Number of task\nruns",
                "duration": "Duration of task\nruns [seconds]",
            }
        )

        row_counter = 0
        rows = dataframe_to_rows(ccsp_report_dataframe, index=False)
        for r_idx, row in enumerate(rows, current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if row_counter == 0:
                    # set header style
                    cell.font = header_font
                    rd = ws.row_dimensions[r_idx]
                    rd.height = 25
                else:
                    # set value style
                    cell.font = value_font

            row_counter += 1

        return current_row + row_counter
